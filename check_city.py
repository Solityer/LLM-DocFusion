import openpyxl
from collections import Counter
import sys

fname = sys.argv[1] if len(sys.argv) > 1 else 'outputs/2025年中国城市经济百强全景报告-模板_filled_f3fae0f5_1.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
rows = [tuple(ws.cell(r,c).value for c in range(1,6)) for r in range(2,ws.max_row+1) if ws.cell(r,1).value]
print(f'Total rows: {len(rows)}')
cities = [r[0] for r in rows]
dupes = {k:v for k,v in Counter(cities).items() if v>1}
print(f'Duplicate cities: {len(dupes)}')
if dupes:
    print('DUPLICATES:', dupes)
else:
    print('NO DUPLICATES - PASS')
print('First 5:', cities[:5])
print('Last 5:', cities[-5:])
