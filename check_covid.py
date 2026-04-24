import openpyxl
from collections import Counter
import sys

fname = sys.argv[1] if len(sys.argv) > 1 else 'outputs/COVID-19 模板_filled_ba30fae3_1.xlsx'
wb = openpyxl.load_workbook(fname)
ws = wb.active
rows = [tuple(ws.cell(r,c).value for c in range(1,7)) for r in range(2,ws.max_row+1) if ws.cell(r,1).value]
print(f'Total rows: {len(rows)}')
countries = [r[0] for r in rows]
# Check for Chinese sub-national regions
chinese_bad = [c for c in countries if c and any(c.endswith(x) for x in ['省','市','自治区','特别行政区'])]
print(f'Chinese sub-national (bad): {len(chinese_bad)}')
if chinese_bad:
    print('  Examples:', chinese_bad[:10])

# Sample rows
print('First 3 rows:')
for r in rows[:3]:
    print(' ', r)
print('Last 3 rows:')
for r in rows[-3:]:
    print(' ', r)

# Date range check
dates = [r[2] for r in rows if r[2]]  # 日期 is col 3
if dates:
    print(f'Date range: {min(dates)} ~ {max(dates)}')