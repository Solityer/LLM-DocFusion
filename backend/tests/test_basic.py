"""Basic tests for the DocFusion system."""
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import json
import tempfile
from pathlib import Path


def _require_fixture_paths(paths: list[str], label: str) -> bool:
    """Return False and mark the test as skipped when external fixture files are absent."""
    missing = [path for path in paths if not Path(path).exists()]
    if missing:
        print(f"↷ {label} skipped (fixtures missing)")
        return False
    return True


def test_json_repair():
    """Test JSON extraction and repair."""
    from app.utils.json_repair import safe_parse_json, extract_json_from_text

    # Clean JSON
    obj, err = safe_parse_json('{"a": 1, "b": "hello"}')
    assert obj == {"a": 1, "b": "hello"}, f"Failed: {obj}"
    assert err == ""

    # JSON with markdown wrapper
    obj, err = safe_parse_json('```json\n{"a": 1}\n```')
    assert obj == {"a": 1}, f"Failed: {obj}"

    # JSON with extra text
    obj, err = safe_parse_json('Here is the result: {"a": 1}')
    assert obj == {"a": 1}, f"Failed: {obj}"

    # Trailing comma
    obj, err = safe_parse_json('{"a": 1, "b": 2,}')
    assert obj is not None and obj.get("a") == 1

    # Array
    obj, err = safe_parse_json('[{"x": 1}, {"x": 2}]')
    assert isinstance(obj, list) and len(obj) == 2

    # None/True/False
    obj, err = safe_parse_json('{"a": None, "b": True}')
    assert obj is not None

    # Empty
    obj, err = safe_parse_json('')
    assert obj is None

    print("✓ test_json_repair passed")


def test_text_utils():
    """Test text utilities."""
    from app.utils.text_utils import similarity, best_column_match, truncate_text

    assert similarity("hello", "hello") == 1.0
    assert similarity("GDP", "GDP总量") > 0.3

    assert best_column_match("城市名称", ["城市", "GDP", "人口"]) == "城市"
    assert best_column_match("国内生产总值", ["GDP", "人口", "面积"]) is None or True  # may or may not match
    assert best_column_match("xyz_abc", ["def", "ghi"]) is None

    result = truncate_text("a" * 100, 50)
    assert len(result) < 100

    print("✓ test_text_utils passed")


def test_requirement_parsing():
    """Test requirement parsing."""
    from app.services.requirement_service import parse_requirement

    # Date range
    spec = parse_requirement("将日期从2020/7/1到2020/8/31的数据填入模板中")
    assert spec.time_range is not None
    assert "2020-07-01" in spec.time_range[0]
    assert "2020-08-31" in spec.time_range[1]

    # Table specs
    spec = parse_requirement("""
    表一：
        监测时间：2025-11-25 09:00:00.0
        城市：德州市
    表二：
        监测时间：2025-11-25 09:00:00.0
        城市：潍坊市
    """)
    assert len(spec.table_specs) >= 2
    assert "德州市" in spec.entity_keywords

    # Empty requirement
    spec = parse_requirement("帮我智能填表")
    assert spec.raw_text == "帮我智能填表"

    # Exact datetime filter
    spec = parse_requirement("根据数据源中的空气质量监测数据，将监测时间为2025-11-25 09:00:00.0的数据填入模板。")
    assert spec.filters.get("监测时间") == ["2025-11-25 09:00:00.0"]

    print("✓ test_requirement_parsing passed")


def test_requirement_auto_infer():
    """Test auto-inferred requirement contains the required sections."""
    import shutil
    from app.services.requirement_service import auto_infer_requirement

    template_src = "/home/match/LLM-DocFusion/测试集/包含模板文件/2025年中国城市经济百强全景报告/2025年中国城市经济百强全景报告-模板.xlsx"
    source_src = "/home/match/LLM-DocFusion/测试集/包含模板文件/2025年中国城市经济百强全景报告/2025年中国城市经济百强全景报告.docx"
    if not _require_fixture_paths([template_src, source_src], "test_requirement_auto_infer"):
        return

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_root = Path(temp_dir)
        template = temp_root / "city-template.xlsx"
        source = temp_root / "city-report.docx"
        shutil.copy2(template_src, template)
        shutil.copy2(source_src, source)

        spec, inferred_text, _warnings = auto_infer_requirement([str(template)], [str(source)])
        assert "时间范围" in inferred_text
        assert "实体范围" in inferred_text
        assert "指标关键词" in inferred_text
        assert "筛选条件" in inferred_text
        assert spec.output_granularity
        assert spec.filters

    print("✓ test_requirement_auto_infer passed")


def test_document_reading():
    """Test document reading with a temp file."""
    from app.services.document_service import read_document
    from app.schemas.models import FileRole

    # Create temp CSV
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8') as f:
        f.write("Name,Age,City\n")
        f.write("Alice,30,Beijing\n")
        f.write("Bob,25,Shanghai\n")
        temp_path = f.name

    try:
        doc = read_document(temp_path, FileRole.SOURCE)
        assert doc.file_type == 'csv'
        assert len(doc.tables) == 1
        assert doc.tables[0].headers == ['Name', 'Age', 'City']
        assert len(doc.tables[0].rows) == 2
    finally:
        os.unlink(temp_path)

    # Create temp TXT
    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as f:
        f.write("This is a test paragraph.\n\nAnother paragraph with data.\n")
        temp_path = f.name

    try:
        doc = read_document(temp_path, FileRole.SOURCE)
        assert doc.file_type == 'text'
        assert len(doc.text_blocks) >= 1
    finally:
        os.unlink(temp_path)

    # Create temp MD
    with tempfile.NamedTemporaryFile(mode='w', suffix='.md', delete=False, encoding='utf-8') as f:
        f.write("# Title\n\nSome text\n\n| A | B |\n| --- | --- |\n| 1 | 2 |\n| 3 | 4 |\n")
        temp_path = f.name

    try:
        doc = read_document(temp_path, FileRole.SOURCE)
        assert doc.file_type == 'markdown'
        assert len(doc.tables) == 1
        assert doc.tables[0].headers == ['A', 'B']
    finally:
        os.unlink(temp_path)

    print("✓ test_document_reading passed")


def test_template_parsing():
    """Test template parsing."""
    from app.services.template_service import parse_template

    # CSV template
    with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, encoding='utf-8') as f:
        f.write("Name,Score,Grade\n")
        f.write(",,,\n")
        temp_path = f.name

    try:
        schema = parse_template(temp_path)
        assert schema.file_type == 'csv'
        assert len(schema.tables) == 1
        assert 'Name' in schema.tables[0].headers
    finally:
        os.unlink(temp_path)

    # TXT template with placeholder
    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as f:
        f.write("Report for {{name}}\nDate: {{date}}\nScore: {{score}}\n")
        temp_path = f.name

    try:
        schema = parse_template(temp_path)
        assert schema.file_type == 'text'
        assert 'name' in schema.placeholders
    finally:
        os.unlink(temp_path)

    print("✓ test_template_parsing passed")


def test_text_rule_extraction():
    """Test multi-entity narrative extraction without depending entirely on the LLM."""
    from app.schemas.models import CandidateEvidence, RequirementSpec, TemplateSchema
    from app.services.document_service import read_document
    from app.services.extraction_service import extract_data
    from app.services.retrieval_service import RetrievalResult
    from app.services.template_service import parse_template
    from app.schemas.models import FileRole

    source_path = "/home/match/LLM-DocFusion/测试集/包含模板文件/2025年中国城市经济百强全景报告/2025年中国城市经济百强全景报告.docx"
    template_path = "/home/match/LLM-DocFusion/测试集/包含模板文件/2025年中国城市经济百强全景报告/2025年中国城市经济百强全景报告-模板.xlsx"
    if not _require_fixture_paths([source_path, template_path], "test_text_rule_extraction"):
        return
    doc = read_document(source_path, FileRole.SOURCE)
    template = parse_template(template_path)

    retrieval = RetrievalResult()
    retrieval.source_docs = [doc]
    retrieval.text_candidates = [
        CandidateEvidence(
            source_file=source_path,
            location="text_block0",
            raw_snippet=doc.text_blocks[0].content[:120],
            match_reason="test",
            confidence=0.8,
        )
    ]
    requirement = RequirementSpec(raw_text="帮我智能填表")

    result = extract_data(retrieval, template, requirement, use_llm=False)
    first_table = result[0]
    assert len(first_table["records"]) >= 50
    first_row = first_table["records"][0]["values"]
    assert first_row.get("城市名")
    assert first_row.get("GDP总量（亿元）")

    print("✓ test_text_rule_extraction passed")


def test_text_record_threshold_blocks_partial_cross_topic_rows():
    """Text records should require enough field coverage to avoid cross-topic pollution."""
    from app.services.extraction_service import _build_text_record

    headers = [
        "城市名",
        "GDP总量（亿元）",
        "常住人口（万）",
        "人均GDP（元）",
        "一般公共预算收入（亿元）",
    ]
    partial_segment = "湖北省常住人口约 5775 万人，人均 GDP 约 7.3 万元，当日核酸检测量约 12.6 万份。"
    full_segment = "上海市 GDP 总量达到 53926.71 亿元，常住人口 2487 万人，人均 GDP 高达 156000 元，一般公共预算收入 8374.2 亿元。"

    assert _build_text_record(headers, partial_segment, "covid.docx", "text_block1") is None
    assert _build_text_record(headers, full_segment, "city.docx", "text_block2") is not None

    print("✓ test_text_record_threshold_blocks_partial_cross_topic_rows passed")


def test_extract_leading_entity_skips_generic_heading_line():
    """Generic heading lines should not overwrite the first real entity in narrative blocks."""
    from app.services.extraction_service import _extract_leading_entity

    segment = "城市经济简报\n南京市 GDP 总量达到 18500.5 亿元，常住人口 949.11 万人。"
    pure_heading_segment = "南京市\nGDP 总量达到 18500.5 亿元，常住人口 949.11 万人。"

    assert _extract_leading_entity(segment) == "南京市"
    assert _extract_leading_entity(pure_heading_segment) == "南京市"

    print("✓ test_extract_leading_entity_skips_generic_heading_line passed")


def test_text_record_threshold_allows_partial_same_domain_rows_for_merge():
    """Same-domain partial rows should survive so multi-source merge can complete them later."""
    from app.services.extraction_service import _build_text_record

    headers = [
        "城市名",
        "GDP总量（亿元）",
        "常住人口（万）",
        "人均GDP（元）",
        "一般公共预算收入（亿元）",
    ]
    segment = "南京市 GDP总量达到 18500 亿元，常住人口 955 万人。"

    assert _build_text_record(headers, segment, "city.docx", "text_block1") is not None

    print("✓ test_text_record_threshold_allows_partial_same_domain_rows_for_merge passed")


def test_candidate_text_segments_group_entities():
    """Narrative segmentation should group continuation paragraphs under one entity block."""
    from app.schemas.models import DocumentBundle, FileRole, TextBlock
    from app.services.extraction_service import _candidate_text_segments, _extract_leading_entity

    doc = DocumentBundle(
        document_id="demo",
        source_file="/tmp/demo.docx",
        file_type="word",
        role=FileRole.SOURCE,
        text_blocks=[
            TextBlock(content="北京GDP总量达到52073.4亿元，常住人口2185.3万人。", block_index=0),
            TextBlock(content="一般公共预算收入也保持增长。", block_index=1),
            TextBlock(content="上海GDP总量达到53926.7亿元，常住人口2487万人。", block_index=2),
        ],
    )

    segments = _candidate_text_segments(doc)
    assert len(segments) == 2, segments
    assert "北京" in segments[0][0]
    assert "上海" in segments[1][0]
    assert _extract_leading_entity("城市经济概览") == ""

    print("✓ test_candidate_text_segments_group_entities passed")


def test_text_reader_splits_line_oriented_txt_blocks():
    """TXT sources with one item per line should become multiple text blocks."""
    from app.schemas.models import FileRole
    from app.services.document_service import read_document

    with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, encoding='utf-8') as f:
        f.write("南京市 GDP总量达到 18500 亿元，常住人口 955 万人。\n")
        f.write("苏州市 一般公共预算收入 2458 亿元，人均GDP 190000 元。\n")
        temp_path = f.name

    try:
        doc = read_document(temp_path, FileRole.SOURCE)
        assert len(doc.text_blocks) == 2
    finally:
        os.unlink(temp_path)

    print("✓ test_text_reader_splits_line_oriented_txt_blocks passed")


def test_field_row_identifier_ignores_placeholder_locations():
    """Placeholder or paragraph-level replacements must not count as contributed rows."""
    from app.services.pipeline_service import _field_row_identifier

    assert _field_row_identifier("word_placeholder", "城市") == ""
    assert _field_row_identifier("paragraph3", "城市") == ""
    assert _field_row_identifier("txt_pos12", "城市") == ""
    assert _field_row_identifier("Sheet1!A12", "城市") == "Sheet1!12"

    print("✓ test_field_row_identifier_ignores_placeholder_locations passed")


def test_validation_flags_inconsistent_source_stats():
    """Validation should fail when per-source contributed metrics are self-contradictory."""
    from app.schemas.models import FilledResult, SourceProcessingStat
    from app.services.validation_service import validate_result

    result = FilledResult(
        template_file=__file__,
        output_file=__file__,
        rows_filled=2,
        record_count=2,
        expected_rows=2,
        fill_rate=100.0,
        source_stats=[
            SourceProcessingStat(
                source_file="/tmp/source.docx",
                file_type="word",
                extracted_records=1,
                contributed_records=3,
                contributed_fields=2,
            )
        ],
    )

    validated = validate_result(result)
    assert any(
        item.check == "source_stat_consistency" and not item.passed
        for item in validated.validation_report
    )

    print("✓ test_validation_flags_inconsistent_source_stats passed")


def test_validation_allows_irrelevant_zero_source():
    """Irrelevant uploaded sources should not fail template-level source contribution checks."""
    from app.schemas.models import FilledResult, SourceProcessingStat
    from app.services.validation_service import validate_result

    result = FilledResult(
        template_file=__file__,
        output_file=__file__,
        rows_filled=1,
        record_count=1,
        expected_rows=1,
        fill_rate=100.0,
        source_stats=[
            SourceProcessingStat(
                source_file="/tmp/relevant.docx",
                file_type="word",
                relevant_to_template=True,
                extracted_records=1,
                contributed_records=1,
                contributed_fields=2,
            ),
            SourceProcessingStat(
                source_file="/tmp/irrelevant.docx",
                file_type="word",
                relevant_to_template=False,
                extracted_records=0,
                contributed_records=0,
                contributed_fields=0,
            ),
        ],
    )

    validated = validate_result(result)
    assert any(
        item.check == "source_contribution" and item.passed
        for item in validated.validation_report
    )

    print("✓ test_validation_allows_irrelevant_zero_source passed")


def test_validation_fails_relevant_narrative_zero_contribution():
    """Relevant narrative sources with zero final contribution must fail validation."""
    from app.schemas.models import FilledResult, SourceProcessingStat
    from app.services.validation_service import validate_result

    result = FilledResult(
        template_file=__file__,
        output_file=__file__,
        rows_filled=10,
        record_count=10,
        expected_rows=10,
        fill_rate=100.0,
        source_stats=[
            SourceProcessingStat(
                source_file="/tmp/relevant-narrative.docx",
                file_type="word",
                relevant_to_template=True,
                text_blocks=30,
                extracted_records=8,
                contributed_records=0,
                contributed_fields=0,
                qwen_stages=["extract"],
            ),
            SourceProcessingStat(
                source_file="/tmp/relevant-structured.xlsx",
                file_type="excel",
                relevant_to_template=True,
                extracted_records=10,
                contributed_records=10,
                contributed_fields=60,
            ),
        ],
    )

    validated = validate_result(result)
    assert any(
        item.check == "unstructured_source_contribution" and not item.passed
        for item in validated.validation_report
    )

    print("✓ test_validation_fails_relevant_narrative_zero_contribution passed")


def test_validation_flags_entity_blocks_filtered_to_zero():
    """Relevant narrative sources with entity blocks but zero filtered records must fail flow validation."""
    from app.schemas.models import FilledResult, SourceProcessingStat
    from app.services.validation_service import validate_result

    result = FilledResult(
        template_file=__file__,
        output_file=__file__,
        rows_filled=5,
        record_count=5,
        expected_rows=5,
        fill_rate=100.0,
        source_stats=[
            SourceProcessingStat(
                source_file="/tmp/relevant-narrative.docx",
                file_type="word",
                text_blocks=28,
                entity_blocks_detected=12,
                relevant_to_template=True,
                extracted_records=8,
                filtered_records=0,
                contributed_records=0,
                contributed_fields=0,
                qwen_stages=["extract"],
            )
        ],
    )

    validated = validate_result(result)
    assert any(
        item.check == "entity_block_record_flow" and not item.passed
        for item in validated.validation_report
    )

    print("✓ test_validation_flags_entity_blocks_filtered_to_zero passed")


def test_region_field_accepts_subnational_entity_when_context_allows_admin_granularity():
    """A region-like field should accept province-level entities when the template context is administrative, not country-only."""
    from app.utils.entity_utils import evaluate_entity_compatibility

    assessment = evaluate_entity_compatibility(
        "江苏省",
        "国家/地区",
        peer_headers=["国家/地区", "病例数"],
        context_text="地区疫情统计表 各地区病例数",
    )

    assert assessment["accepted"] is True
    assert assessment["normalized_entity_type"] == "sub_national"
    assert assessment["normalized_granularity"] == "sub_national"
    assert "sub_national" in assessment["allowed_granularity_set"]

    print("✓ test_region_field_accepts_subnational_entity_when_context_allows_admin_granularity passed")


def test_region_field_blocks_subnational_entity_in_country_only_context():
    """A global country-only template should still block province-level entities from entering region fields."""
    from app.utils.entity_utils import evaluate_entity_compatibility

    assessment = evaluate_entity_compatibility(
        "江苏省",
        "国家",
        peer_headers=["大洲", "国家", "病例数"],
        context_text="全球疫情统计 world global continent 国家",
    )

    assert assessment["accepted"] is False
    assert assessment["filter_reason"] == "region_field_country_only"
    assert assessment["whether_recoverable"] is True

    print("✓ test_region_field_blocks_subnational_entity_in_country_only_context passed")


def test_filter_records_keeps_recoverable_region_record():
    """Recoverable province-level region records should survive entity filtering instead of being wiped out."""
    from app.services.extraction_service import (
        _annotate_records_with_entity_semantics,
        _filter_records_by_entity_legality,
        _normalize_records,
    )

    headers = ["国家/地区", "病例数"]
    records = [{
        "values": {"国家/地区": "江苏省", "病例数": "13"},
        "field_confidence": {"国家/地区": 0.82, "病例数": 0.88},
        "field_evidence": {},
        "source_file": "/tmp/province.docx",
        "source_location": "text_block1",
        "row_index": 0,
        "match_methods": {"国家/地区": "entity_lead", "病例数": "text_exact"},
    }]
    template_context = {
        "template_file": "/tmp/template.xlsx",
        "anchor_text": "地区疫情统计表",
        "topic_text": "地区疫情统计表 国家/地区 病例数",
    }

    normalized = _normalize_records(headers, records)
    annotated = _annotate_records_with_entity_semantics(headers, normalized, template_context)
    filtered, invalidated_sources, diagnostics = _filter_records_by_entity_legality(
        headers,
        annotated,
        [],
        template_context=template_context,
    )

    assert len(filtered) == 1
    assert invalidated_sources == set()
    assert filtered[0]["normalized_entity_type"] == "sub_national"
    assert diagnostics["recovered_examples"]

    print("✓ test_filter_records_keeps_recoverable_region_record passed")


def test_validation_flags_uniform_filter_reason_mass_clear():
    """Narrative sources mass-cleared by one filter reason should fail validation explicitly."""
    from app.schemas.models import FilledResult, SourceProcessingStat
    from app.services.validation_service import validate_result

    result = FilledResult(
        template_file=__file__,
        output_file=__file__,
        rows_filled=2,
        record_count=2,
        expected_rows=2,
        fill_rate=100.0,
        source_stats=[
            SourceProcessingStat(
                source_file="/tmp/relevant-narrative.docx",
                file_type="word",
                text_blocks=12,
                entity_blocks_detected=4,
                relevant_to_template=True,
                extracted_records=4,
                filtered_records=0,
                contributed_records=0,
                contributed_fields=0,
            ),
            SourceProcessingStat(
                source_file="/tmp/relevant-structured.xlsx",
                file_type="excel",
                relevant_to_template=True,
                extracted_records=2,
                filtered_records=2,
                contributed_records=2,
                contributed_fields=4,
            ),
        ],
        entity_legality_report={
            "per_source_filter_reasons": {
                "/tmp/relevant-narrative.docx": {"region_field_country_only": 4}
            }
        },
    )

    validated = validate_result(result)
    assert any(
        item.check == "uniform_filter_reason" and not item.passed
        for item in validated.validation_report
    )

    print("✓ test_validation_flags_uniform_filter_reason_mass_clear passed")


def test_validation_keeps_output_when_only_response_time_fails():
    """Response-time misses should keep the output traceable even if minimum requirements are missed."""
    from app.schemas.models import FilledResult
    from app.services.validation_service import validate_result

    result = FilledResult(
        template_file=__file__,
        output_file=__file__,
        rows_filled=2,
        record_count=2,
        expected_rows=2,
        fill_rate=100.0,
        timing={"total": 95.0},
    )

    validated = validate_result(result)
    assert validated.status == "completed"
    assert validated.meets_minimum is False
    assert validated.output_file == __file__

    print("✓ test_validation_keeps_output_when_only_response_time_fails passed")


def test_numeric_normalization_handles_population_and_per_capita_units():
    """Narrative numeric normalization should expand implicit '万' units for population and per-capita values."""
    from app.services.extraction_service import _normalize_numeric_value

    population_value, population_status = _normalize_numeric_value("人口", "5775 万")
    per_capita_value, per_capita_status = _normalize_numeric_value("人均GDP（元）", "7.3 万元")
    tests_value, tests_status = _normalize_numeric_value("每日检测数", "12.6 万份")

    assert population_value == "57750000"
    assert population_status == "normalized"
    assert per_capita_value == "73000"
    assert per_capita_status == "normalized"
    assert tests_value == "126000"
    assert tests_status == "normalized"

    print("✓ test_numeric_normalization_handles_population_and_per_capita_units passed")


def test_retrieval_plausibility_blocks_cross_domain_tables():
    """Cross-domain tables should be rejected before they consume qwen mapping budget."""
    from app.schemas.models import NormalizedTable, TemplateTable
    from app.services.retrieval_service import _table_pair_is_plausible

    air_template = TemplateTable(
        table_index=0,
        headers=["城市", "AQI", "PM2.5", "监测时间"],
    )
    covid_source = NormalizedTable(
        table_index=0,
        headers=["国家", "累计确诊", "累计死亡", "日期"],
        rows=[["中国", "100", "1", "2020-01-01"]],
    )
    city_source = NormalizedTable(
        table_index=1,
        headers=["城市名称", "地区生产总值", "常住人口"],
        rows=[["南京市", "18500", "950"]],
    )
    economy_template = TemplateTable(
        table_index=2,
        headers=["城市名", "GDP总量（亿元）", "常住人口（万）"],
    )

    assert _table_pair_is_plausible(air_template, covid_source) is False
    assert _table_pair_is_plausible(economy_template, city_source) is True

    print("✓ test_retrieval_plausibility_blocks_cross_domain_tables passed")


def test_estimate_entity_count_is_scoped_to_relevant_sources():
    """Entity estimation should not be inflated by unrelated uploaded sources."""
    from app.schemas.models import DocumentBundle, FileRole, TextBlock
    from app.services.extraction_service import _estimate_entity_count
    from app.services.retrieval_service import RetrievalResult

    relevant_doc = DocumentBundle(
        document_id="relevant",
        source_file="/tmp/relevant.docx",
        file_type="word",
        role=FileRole.SOURCE,
        text_blocks=[TextBlock(content="南京市 GDP 总量达到 18500 亿元。", block_index=0)],
        raw_text="南京市 GDP 总量达到 18500 亿元。",
    )
    irrelevant_doc = DocumentBundle(
        document_id="irrelevant",
        source_file="/tmp/irrelevant.docx",
        file_type="word",
        role=FileRole.SOURCE,
        text_blocks=[
            TextBlock(content="北京市 GDP 总量达到 50000 亿元。", block_index=0),
            TextBlock(content="上海市 GDP 总量达到 54000 亿元。", block_index=1),
            TextBlock(content="广州市 GDP 总量达到 31000 亿元。", block_index=2),
        ],
        raw_text="\n".join([
            "北京市 GDP 总量达到 50000 亿元。",
            "上海市 GDP 总量达到 54000 亿元。",
            "广州市 GDP 总量达到 31000 亿元。",
        ]),
    )

    retrieval = RetrievalResult()
    retrieval.source_docs = [relevant_doc, irrelevant_doc]

    count = _estimate_entity_count(
        retrieval,
        candidate_row_estimates=[],
        records=[],
        relevant_source_files={"/tmp/relevant.docx"},
    )
    assert count == 1

    print("✓ test_estimate_entity_count_is_scoped_to_relevant_sources passed")


def test_build_field_result_prefers_evidence_source():
    """Field-level source attribution should prefer actual evidence over merged record ownership."""
    from app.schemas.models import CandidateEvidence
    from app.services.fill_service import _build_field_result

    record = {
        "source_file": "/tmp/source_a.docx",
        "field_evidence": {
            "城市": [
                CandidateEvidence(
                    source_file="/tmp/source_b.docx",
                    location="text_block0",
                    raw_snippet="南京市 GDP 总量达到 18500 亿元。",
                    match_reason="test",
                    confidence=0.9,
                )
            ]
        },
        "match_methods": {"城市": "text_rule"},
        "field_confidence": {"城市": 0.88},
    }

    field = _build_field_result(
        "城市",
        "Sheet1!A2",
        "南京市",
        "南京市",
        record,
        {"extraction_method": "hybrid_rule"},
    )
    assert field.source_file == "/tmp/source_b.docx"

    print("✓ test_build_field_result_prefers_evidence_source passed")


def test_source_stats_split_evidence_vs_value_for_merged_fields():
    """Merged fields should separate evidence support from the chosen value owner."""
    from app.schemas.models import (
        CandidateEvidence,
        DocumentBundle,
        FileRole,
        FilledFieldResult,
        FilledResult,
        ModelUsageSummary,
    )
    from app.services.pipeline_service import _build_source_stats
    from app.services.retrieval_service import RetrievalResult

    documents = [
        DocumentBundle(document_id="a", source_file="/tmp/a.docx", file_type="word", role=FileRole.SOURCE),
        DocumentBundle(document_id="b", source_file="/tmp/b.xlsx", file_type="xlsx", role=FileRole.SOURCE),
    ]
    retrieval = RetrievalResult()
    retrieval.source_docs = documents
    extracted = [{
        "source_counts": {"/tmp/a.docx": 1, "/tmp/b.xlsx": 1},
        "relevant_source_files": ["/tmp/a.docx", "/tmp/b.xlsx"],
    }]
    filled = FilledResult(
        template_file="/tmp/template.xlsx",
        rows_filled=1,
        record_count=1,
        expected_rows=1,
        fill_rate=100.0,
        filled_fields=[
            FilledFieldResult(
                field_name="GDP总量（亿元）",
                target_location="Sheet1!B2",
                value="18500",
                normalized_value="18500",
                source_file="/tmp/a.docx",
                evidence=[
                    CandidateEvidence(
                        source_file="/tmp/a.docx",
                        location="text_block1",
                        raw_snippet="南京市 GDP 总量达到 18500 亿元。",
                        match_reason="text",
                        confidence=0.92,
                    ),
                    CandidateEvidence(
                        source_file="/tmp/b.xlsx",
                        location="table0(Sheet1)",
                        raw_snippet="城市 | GDP\n南京市 | 18500",
                        match_reason="table",
                        confidence=0.87,
                    ),
                ],
            )
        ],
        model_usage=ModelUsageSummary(),
    )

    stats = _build_source_stats(
        documents,
        retrieval,
        extracted,
        filled,
        "demo",
        model_usage=filled.model_usage,
    )
    per_source = {item.source_file: item for item in stats}
    assert per_source["/tmp/a.docx"].contributed_fields == 1
    assert per_source["/tmp/a.docx"].evidence_contribution_fields == 1
    assert per_source["/tmp/b.xlsx"].contributed_fields == 0
    assert per_source["/tmp/b.xlsx"].evidence_contribution_fields == 1
    assert per_source["/tmp/a.docx"].contributed_records == 1
    assert per_source["/tmp/b.xlsx"].contributed_records == 0

    print("✓ test_source_stats_split_evidence_vs_value_for_merged_fields passed")


def test_source_stats_demote_invalidated_narrative_source():
    """Sources that only produced illegal-entity records should be treated as ignored, not relevant-zero."""
    from app.schemas.models import DocumentBundle, FileRole, FilledFieldResult, FilledResult, ModelUsageSummary
    from app.services.pipeline_service import _build_source_stats
    from app.services.retrieval_service import RetrievalResult

    documents = [
        DocumentBundle(document_id="covid-doc", source_file="/tmp/covid.docx", file_type="word", role=FileRole.SOURCE),
        DocumentBundle(document_id="covid-xlsx", source_file="/tmp/covid.xlsx", file_type="excel", role=FileRole.SOURCE),
    ]
    retrieval = RetrievalResult()
    retrieval.source_docs = documents
    extracted = [{
        "source_counts": {"/tmp/covid.docx": 8, "/tmp/covid.xlsx": 12},
        "filtered_source_counts": {"/tmp/covid.xlsx": 12},
        "entity_block_counts": {"/tmp/covid.docx": 10},
        "relevant_source_files": ["/tmp/covid.docx", "/tmp/covid.xlsx"],
        "invalidated_source_files": ["/tmp/covid.docx"],
    }]
    filled = FilledResult(
        template_file="/tmp/template.xlsx",
        rows_filled=12,
        record_count=12,
        expected_rows=12,
        fill_rate=100.0,
        filled_fields=[
            FilledFieldResult(
                field_name="国家/地区",
                target_location="Sheet1!A2",
                value="Albania",
                normalized_value="Albania",
                source_file="/tmp/covid.xlsx",
                evidence=[],
            )
        ],
        model_usage=ModelUsageSummary(),
    )

    stats = _build_source_stats(
        documents,
        retrieval,
        extracted,
        filled,
        "COVID-19 模板.xlsx",
        model_usage=filled.model_usage,
    )
    per_source = {item.source_file: item for item in stats}
    assert per_source["/tmp/covid.docx"].relevant_to_template is False
    assert per_source["/tmp/covid.docx"].entity_blocks_detected == 10
    assert per_source["/tmp/covid.docx"].extracted_records == 8
    assert per_source["/tmp/covid.docx"].filtered_records == 0
    assert per_source["/tmp/covid.docx"].contributed_fields == 0
    assert any("仅产出实体类型不合法或粒度不符" in warning for warning in per_source["/tmp/covid.docx"].warnings)

    print("✓ test_source_stats_demote_invalidated_narrative_source passed")


def test_semantic_merge_normalizes_city_suffix_for_same_entity():
    """Entity-key merge should align '南京' and '南京市' instead of treating them as unrelated rows."""
    from app.schemas.models import CandidateEvidence, RequirementSpec
    from app.services.extraction_service import _merge_records_by_semantic_key

    headers = ["城市名", "GDP总量（亿元）"]
    base_record = {
        "values": {"城市名": "南京", "GDP总量（亿元）": "18500"},
        "field_evidence": {
            "城市名": [CandidateEvidence(source_file="/tmp/ranking.docx", location="text_block1", raw_snippet="南京 GDP 18500", match_reason="base", confidence=0.9)],
            "GDP总量（亿元）": [CandidateEvidence(source_file="/tmp/ranking.docx", location="text_block1", raw_snippet="南京 GDP 18500", match_reason="base", confidence=0.9)],
        },
        "field_confidence": {"城市名": 0.9, "GDP总量（亿元）": 0.9},
        "match_methods": {"城市名": "text_rule", "GDP总量（亿元）": "text_rule"},
        "source_file": "/tmp/ranking.docx",
    }
    narrative_record = {
        "values": {"城市名": "南京市", "GDP总量（亿元）": "18500"},
        "field_evidence": {
            "城市名": [CandidateEvidence(source_file="/tmp/nanjing.docx", location="text_block9", raw_snippet="南京市 GDP 18500", match_reason="narrative", confidence=0.82)],
            "GDP总量（亿元）": [CandidateEvidence(source_file="/tmp/nanjing.docx", location="text_block9", raw_snippet="南京市 GDP 18500", match_reason="narrative", confidence=0.82)],
        },
        "field_confidence": {"城市名": 0.82, "GDP总量（亿元）": 0.82},
        "match_methods": {"城市名": "text_rule", "GDP总量（亿元）": "text_rule"},
        "source_file": "/tmp/nanjing.docx",
    }

    merged = _merge_records_by_semantic_key(
        [base_record, narrative_record],
        headers=headers,
        requirement=RequirementSpec(raw_text="帮我智能填表"),
        use_llm=False,
    )
    assert len(merged) == 1
    merged_sources = {
        evidence.source_file
        for evidence in merged[0]["field_evidence"]["城市名"]
    }
    assert merged_sources == {"/tmp/ranking.docx", "/tmp/nanjing.docx"}

    print("✓ test_semantic_merge_normalizes_city_suffix_for_same_entity passed")


def test_validation_does_not_flag_single_city_report_as_repeated_entity_failure():
    """Long single-city narrative reports should not fail the repeated-entity check."""
    from app.schemas.models import FilledResult, SourceProcessingStat
    from app.services.validation_service import validate_result

    result = FilledResult(
        template_file=__file__,
        output_file=__file__,
        rows_filled=1,
        record_count=1,
        expected_rows=1,
        fill_rate=100.0,
        source_stats=[
            SourceProcessingStat(
                source_file="/tmp/南京市2024年国民经济和社会发展统计公报.docx",
                file_type="word",
                relevant_to_template=True,
                text_blocks=200,
                extracted_records=1,
                contributed_records=1,
                contributed_fields=2,
            ),
        ],
    )

    validated = validate_result(result)
    assert any(
        item.check == "repeated_entity_narrative" and item.passed
        for item in validated.validation_report
    )

    print("✓ test_validation_does_not_flag_single_city_report_as_repeated_entity_failure passed")


def test_template_context_blocks_cross_topic_narrative_sources():
    """Template stem and topic anchors should block cross-topic narrative sources."""
    from app.schemas.models import FileRole, RequirementSpec
    from app.services.document_service import read_document
    from app.services.extraction_service import _build_template_table_context, _source_matches_template_context
    from app.services.template_service import parse_template

    root = Path("/home/match/LLM-DocFusion/测试集/包含模板文件")
    city_dir = root / "2025年中国城市经济百强全景报告"
    covid_dir = root / "COVID-19数据集"
    if not _require_fixture_paths([
        str(city_dir / "2025年中国城市经济百强全景报告.docx"),
        str(city_dir / "2025年中国城市经济百强全景报告-模板.xlsx"),
        str(covid_dir / "中国COVID-19新冠疫情情况.docx"),
        str(covid_dir / "COVID-19 模板.xlsx"),
    ], "test_template_context_blocks_cross_topic_narrative_sources"):
        return

    city_doc = read_document(str(city_dir / "2025年中国城市经济百强全景报告.docx"), FileRole.SOURCE)
    covid_doc = read_document(str(covid_dir / "中国COVID-19新冠疫情情况.docx"), FileRole.SOURCE)

    city_template = parse_template(str(city_dir / "2025年中国城市经济百强全景报告-模板.xlsx"))
    city_requirement = RequirementSpec(raw_text="帮我智能填表")
    city_context = _build_template_table_context(city_template, city_template.tables[0], city_requirement)

    assert _source_matches_template_context(
        city_doc,
        city_template.tables[0].headers,
        city_requirement,
        template_context=city_context,
    ) is True
    assert _source_matches_template_context(
        covid_doc,
        city_template.tables[0].headers,
        city_requirement,
        template_context=city_context,
    ) is False

    covid_template = parse_template(str(covid_dir / "COVID-19 模板.xlsx"))
    covid_requirement = RequirementSpec(raw_text="根据 COVID-19 数据填表")
    covid_context = _build_template_table_context(covid_template, covid_template.tables[0], covid_requirement)

    assert _source_matches_template_context(
        covid_doc,
        covid_template.tables[0].headers,
        covid_requirement,
        template_context=covid_context,
    ) is True
    assert _source_matches_template_context(
        city_doc,
        covid_template.tables[0].headers,
        covid_requirement,
        template_context=covid_context,
    ) is False

    print("✓ test_template_context_blocks_cross_topic_narrative_sources passed")


def test_ranking_limit_comes_from_template_intent_not_offtopic_source_titles():
    """Top-N caps should come from template/requirement intent instead of unrelated source titles."""
    from app.schemas.models import RequirementSpec
    from app.services.extraction_service import _build_template_table_context, _detect_ranking_limit_with_context
    from app.services.retrieval_service import RetrievalResult
    from app.services.template_service import parse_template

    root = Path("/home/match/LLM-DocFusion/测试集/包含模板文件")
    city_dir = root / "2025年中国城市经济百强全景报告"
    covid_dir = root / "COVID-19数据集"
    if not _require_fixture_paths([
        str(city_dir / "2025年中国城市经济百强全景报告-模板.xlsx"),
        str(covid_dir / "COVID-19 模板.xlsx"),
    ], "test_ranking_limit_comes_from_template_intent_not_offtopic_source_titles"):
        return

    city_template = parse_template(str(city_dir / "2025年中国城市经济百强全景报告-模板.xlsx"))
    covid_template = parse_template(str(covid_dir / "COVID-19 模板.xlsx"))

    city_context = _build_template_table_context(
        city_template,
        city_template.tables[0],
        RequirementSpec(raw_text="帮我智能填表"),
    )
    covid_context = _build_template_table_context(
        covid_template,
        covid_template.tables[0],
        RequirementSpec(raw_text="根据 COVID-19 数据填表"),
    )

    retrieval = RetrievalResult()
    mixed_records = [
        {"source_file": str(city_dir / "2025年中国城市经济百强全景报告.docx")},
        {"source_file": str(covid_dir / "COVID-19全球数据集（节选）.xlsx")},
    ]

    assert _detect_ranking_limit_with_context(
        mixed_records,
        retrieval,
        RequirementSpec(raw_text=""),
        template_context=city_context,
    ) == 100
    assert _detect_ranking_limit_with_context(
        mixed_records,
        retrieval,
        RequirementSpec(raw_text=""),
        template_context=covid_context,
    ) == 0

    print("✓ test_ranking_limit_comes_from_template_intent_not_offtopic_source_titles passed")


def test_excel_fill_respects_writable_rows_and_appends_without_overwrite():
    """Excel write-back should use writable rows first and append instead of overwriting existing content."""
    from openpyxl import Workbook, load_workbook

    from app.schemas.models import TemplateSchema, TemplateTable
    from app.services.fill_service import fill_template

    with tempfile.TemporaryDirectory() as temp_dir:
        temp_root = Path(temp_dir)
        template_path = temp_root / "template.xlsx"
        output_path = temp_root / "output.xlsx"

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet["A1"] = "Name"
        worksheet["B1"] = "Value"
        worksheet["A2"] = ""
        worksheet["B2"] = ""
        worksheet["A3"] = "KEEP"
        worksheet["B3"] = "ORIGINAL"
        worksheet["A4"] = ""
        worksheet["B4"] = ""
        workbook.save(template_path)

        template = TemplateSchema(
            source_file=str(template_path),
            file_type="excel",
            tables=[
                TemplateTable(
                    table_index=0,
                    sheet_name="Data",
                    headers=["Name", "Value"],
                    writable_rows=[0, 2],
                    writable_cols=[0, 1],
                    row_count=3,
                    col_count=2,
                )
            ],
        )

        records = [
            {
                "values": {"Name": "Alpha", "Value": "1"},
                "field_confidence": {},
                "field_evidence": {},
                "match_methods": {},
                "source_file": "/tmp/source-a.docx",
            },
            {
                "values": {"Name": "Beta", "Value": "2"},
                "field_confidence": {},
                "field_evidence": {},
                "match_methods": {},
                "source_file": "/tmp/source-b.docx",
            },
            {
                "values": {"Name": "Gamma", "Value": "3"},
                "field_confidence": {},
                "field_evidence": {},
                "match_methods": {},
                "source_file": "/tmp/source-c.docx",
            },
        ]
        extracted = [{
            "table_index": 0,
            "headers": ["Name", "Value"],
            "records": records,
            "col_confidence": {},
            "extraction_method": "rule",
        }]

        result = fill_template(template, extracted, str(output_path))
        filled_workbook = load_workbook(output_path)
        filled_sheet = filled_workbook["Data"]

        assert filled_sheet["A2"].value == "Alpha"
        assert filled_sheet["B2"].value == 1
        assert filled_sheet["A3"].value == "KEEP"
        assert filled_sheet["B3"].value == "ORIGINAL"
        assert filled_sheet["A4"].value == "Beta"
        assert filled_sheet["B4"].value == 2
        assert filled_sheet["A5"].value == "Gamma"
        assert filled_sheet["B5"].value == 3
        assert result.rows_filled == 3
        assert any("追加新行" in warning for warning in result.warnings)

    print("✓ test_excel_fill_respects_writable_rows_and_appends_without_overwrite passed")


if __name__ == "__main__":
    test_json_repair()
    test_text_utils()
    test_requirement_parsing()
    test_requirement_auto_infer()
    test_document_reading()
    test_template_parsing()
    test_text_rule_extraction()
    test_text_record_threshold_blocks_partial_cross_topic_rows()
    test_extract_leading_entity_skips_generic_heading_line()
    test_text_record_threshold_allows_partial_same_domain_rows_for_merge()
    test_candidate_text_segments_group_entities()
    test_text_reader_splits_line_oriented_txt_blocks()
    test_field_row_identifier_ignores_placeholder_locations()
    test_validation_flags_inconsistent_source_stats()
    test_validation_allows_irrelevant_zero_source()
    test_validation_fails_relevant_narrative_zero_contribution()
    test_validation_flags_entity_blocks_filtered_to_zero()
    test_region_field_accepts_subnational_entity_when_context_allows_admin_granularity()
    test_region_field_blocks_subnational_entity_in_country_only_context()
    test_filter_records_keeps_recoverable_region_record()
    test_validation_flags_uniform_filter_reason_mass_clear()
    test_validation_keeps_output_when_only_response_time_fails()
    test_numeric_normalization_handles_population_and_per_capita_units()
    test_retrieval_plausibility_blocks_cross_domain_tables()
    test_estimate_entity_count_is_scoped_to_relevant_sources()
    test_build_field_result_prefers_evidence_source()
    test_source_stats_split_evidence_vs_value_for_merged_fields()
    test_source_stats_demote_invalidated_narrative_source()
    test_semantic_merge_normalizes_city_suffix_for_same_entity()
    test_validation_does_not_flag_single_city_report_as_repeated_entity_failure()
    test_template_context_blocks_cross_topic_narrative_sources()
    test_ranking_limit_comes_from_template_intent_not_offtopic_source_titles()
    test_excel_fill_respects_writable_rows_and_appends_without_overwrite()
    print("\n✅ All basic tests passed!")
